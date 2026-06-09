# -*- coding: utf-8 -*-
"""listings.json + package_details.json + lecture_details.json
   -> 카테고리별 Excel/CSV DB (강좌목록 + 패키지구성 + 전체 차시목차)."""
import sys, json, re, csv
sys.stdout.reconfigure(encoding='utf-8')
from collections import Counter
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = 'https://www.itsdong.com/it/'
data = json.load(open('listings.json', encoding='utf-8'))
pkgdet = json.load(open('package_details.json', encoding='utf-8'))
try:
    lecdet = json.load(open('lecture_details.json', encoding='utf-8'))
except FileNotFoundError:
    lecdet = {}
lectures = data['lectures']
packages = data['packages']

MAIN_ORDER = ['프로그래밍', '그래픽', '컴퓨터일반', '자격증']
SUB_ORDER = {}
for lec in lectures:
    key = (lec['main_name'], lec['sub_name'])
    if key not in SUB_ORDER:
        SUB_ORDER[key] = len(SUB_ORDER)


def won(s):
    if not s:
        return None
    digits = re.sub(r'[^\d]', '', s)
    return int(digits) if digits else None


def pct(s):
    if not s:
        return None
    m = re.search(r'(\d+)', s)
    return int(m.group(1)) if m else None


def norm_level(lv):
    if not lv:
        return None
    m = {'초': '초급', '초급': '초급', '기초': '초급', '하': '초급',
         '중': '중급', '중급': '중급', '중상': '중급',
         '고': '고급', '고급': '고급',
         '초중급': '초중급', '초/중': '초중급', '초/중급': '초중급',
         '중고급': '중고급', '중/고급': '중고급', '활용': '활용'}
    return m.get(lv, lv)


def det(lid):
    return lecdet.get(str(lid), {})


# ---------- 스타일 ----------
HDR_FILL = PatternFill('solid', fgColor='2F3A4A')
HDR_FONT = Font(bold=True, color='FFFFFF', size=10)
CAT_FILL = PatternFill('solid', fgColor='E8ECF1')
LINK_FONT = Font(color='1A56DB', underline='single')
CEN = Alignment(horizontal='center', vertical='center')


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CEN
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = 'A1:%s1' % get_column_letter(ncols)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


wb = openpyxl.Workbook()
wb.remove(wb.active)

# ================= 1) 요약 =================
ws = wb.create_sheet('요약')
ws['A1'] = '아이티동스쿨(itsdong.com) 강좌 DB'
ws['A1'].font = Font(bold=True, size=14)
rows0 = [('수집일', '2026-06-09'),
         ('단일강좌(lecture) 총계', len(lectures)),
         ('패키지(package) 총계', len(packages)),
         ('총 차시(개별 강의) 수', sum(det(l['id']).get('lesson_count', 0) for l in lectures))]
for i, (k, v) in enumerate(rows0, 3):
    ws.cell(row=i, column=1, value=k).font = Font(bold=True)
    ws.cell(row=i, column=2, value=v)
ws['A8'] = '── 대분류 / 소분류별 강좌 수 ──'; ws['A8'].font = Font(bold=True)
hr = 9
for c, t in enumerate(['대분류', '소분류', '강좌수', '총차시수'], 1):
    cell = ws.cell(row=hr, column=c, value=t)
    cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CEN
row = hr + 1
counts = Counter((l['main_name'], l['sub_name']) for l in lectures)
gang = Counter()
for l in lectures:
    gang[(l['main_name'], l['sub_name'])] += det(l['id']).get('lesson_count', 0)
for main in MAIN_ORDER:
    subs = sorted([k for k in counts if k[0] == main], key=lambda k: SUB_ORDER.get(k, 999))
    st = sg = 0
    for k in subs:
        ws.cell(row=row, column=1, value=k[0])
        ws.cell(row=row, column=2, value=k[1])
        ws.cell(row=row, column=3, value=counts[k]).alignment = CEN
        ws.cell(row=row, column=4, value=gang[k]).alignment = CEN
        st += counts[k]; sg += gang[k]; row += 1
    c2 = ws.cell(row=row, column=2, value='▶ %s 소계' % main); c2.font = Font(bold=True)
    for col, val in ((3, st), (4, sg)):
        cc = ws.cell(row=row, column=col, value=val); cc.font = Font(bold=True); cc.alignment = CEN
    for c in range(1, 5):
        ws.cell(row=row, column=c).fill = CAT_FILL
    row += 1
set_widths(ws, [16, 34, 10, 10])

# ================= 2) 대분류별 강좌 =================
LEC_HDR = ['소분류', 'ID', '제목', '강사', '강의수', '난이도', '수강기간',
           '가격(원)', '정가(원)', '할인%', '구분', 'URL']
for main in MAIN_ORDER:
    ws = wb.create_sheet(main)
    ws.append(LEC_HDR)
    rows = [l for l in lectures if l['main_name'] == main]
    rows.sort(key=lambda l: (SUB_ORDER.get((l['main_name'], l['sub_name']), 999), l['title']))
    for l in rows:
        d = det(l['id'])
        url = BASE + 'lecture_detail.php?id=%s' % l['id']
        ws.append([l['sub_name'], int(l['id']), l['title'], l.get('instructor'),
                   d.get('lesson_count'), norm_level(d.get('level')), l.get('period'),
                   won(l.get('price')), won(l.get('list_price')), pct(l.get('discount')),
                   '단일강좌', url])
        r = ws.max_row
        ws.cell(row=r, column=2).alignment = CEN
        ws.cell(row=r, column=5).alignment = CEN
        ws.cell(row=r, column=8).number_format = '#,##0'
        ws.cell(row=r, column=9).number_format = '#,##0'
        lk = ws.cell(row=r, column=12); lk.hyperlink = url; lk.font = LINK_FONT
    style_header(ws, len(LEC_HDR))
    set_widths(ws, [22, 7, 58, 11, 7, 8, 9, 10, 10, 7, 9, 44])

# ================= 3) 패키지 =================
ws = wb.create_sheet('패키지')
PKG_HDR = ['대분류', '대표소분류', 'ID', '패키지명', '수강기간', '가격(원)', '정가(원)',
           '할인%', '포함트랙수', '구분', 'URL']
ws.append(PKG_HDR)
pkgs_sorted = sorted(packages, key=lambda p: (MAIN_ORDER.index(p['main_name'])
                     if p['main_name'] in MAIN_ORDER else 9, -(won(p.get('price')) or 0)))
for p in pkgs_sorted:
    dp = pkgdet.get(p['id'], {})
    url = BASE + 'package_detail.php?id=%s' % p['id']
    ws.append([p['main_name'], p['sub_name'], int(p['id']), dp.get('title') or p['title'],
               p.get('period'), won(p.get('price')), won(p.get('list_price')),
               pct(p.get('discount')), dp.get('track_count'), '패키지', url])
    r = ws.max_row
    ws.cell(row=r, column=3).alignment = CEN
    ws.cell(row=r, column=6).number_format = '#,##0'
    ws.cell(row=r, column=7).number_format = '#,##0'
    ws.cell(row=r, column=9).alignment = CEN
    lk = ws.cell(row=r, column=11); lk.hyperlink = url; lk.font = LINK_FONT
style_header(ws, len(PKG_HDR))
set_widths(ws, [12, 22, 7, 50, 10, 11, 11, 7, 11, 9, 44])

# ================= 4) 패키지 구성(트랙) =================
ws = wb.create_sheet('패키지_구성')
TRK_HDR = ['패키지ID', '패키지명', '순번', '트랙ID', '트랙(강좌)제목', '트랙차시수', '트랙URL']
ws.append(TRK_HDR)
for p in pkgs_sorted:
    dp = pkgdet.get(p['id'], {})
    pname = dp.get('title') or p['title']
    for i, t in enumerate(dp.get('tracks', []), 1):
        turl = BASE + 'lecture_detail.php?id=%s' % t['id']
        ws.append([int(p['id']), pname, i, int(t['id']), t['title'],
                   det(t['id']).get('lesson_count'), turl])
        r = ws.max_row
        lk = ws.cell(row=r, column=7); lk.hyperlink = turl; lk.font = LINK_FONT
style_header(ws, len(TRK_HDR))
set_widths(ws, [9, 44, 6, 7, 56, 9, 44])

# ================= 5) 전체 차시 목차 =================
ws = wb.create_sheet('차시목차')
CUR_HDR = ['대분류', '소분류', '강좌ID', '강좌제목', '차시', '차시(강의)제목', '시간']
ws.append(CUR_HDR)
lec_by_id = {l['id']: l for l in lectures}
for l in sorted(lectures, key=lambda l: (MAIN_ORDER.index(l['main_name']) if l['main_name'] in MAIN_ORDER else 9,
                                         SUB_ORDER.get((l['main_name'], l['sub_name']), 999), l['title'])):
    d = det(l['id'])
    for les in d.get('lessons', []):
        ws.append([l['main_name'], l['sub_name'], int(l['id']), l['title'],
                   les['no'], les['title'], les.get('time')])
style_header(ws, len(CUR_HDR))
set_widths(ws, [12, 22, 7, 50, 6, 60, 9])

OUT = 'itsdong_강좌DB.xlsx'
wb.save(OUT)
print('저장:', OUT, '(차시목차 %d행)' % (ws.max_row - 1))

# ================= CSV =================
def write_csv(fn, header, rows):
    with open(fn, 'w', encoding='utf-8-sig', newline='') as f:
        w = csv.writer(f); w.writerow(header); w.writerows(rows)

lec_rows = []
for l in sorted(lectures, key=lambda l: (MAIN_ORDER.index(l['main_name']) if l['main_name'] in MAIN_ORDER else 9,
                                         SUB_ORDER.get((l['main_name'], l['sub_name']), 999), l['title'])):
    d = det(l['id'])
    lec_rows.append([l['main_name'], l['sub_name'], '단일강좌', l['id'], l['title'],
                     l.get('instructor'), d.get('lesson_count'), norm_level(d.get('level')),
                     l.get('period'), won(l.get('price')), won(l.get('list_price')),
                     pct(l.get('discount')), BASE + 'lecture_detail.php?id=%s' % l['id']])
write_csv('itsdong_강좌목록.csv',
          ['대분류', '소분류', '구분', 'ID', '제목', '강사', '강의수', '난이도', '수강기간',
           '가격', '정가', '할인%', 'URL'], lec_rows)

pkg_rows = []
for p in pkgs_sorted:
    dp = pkgdet.get(p['id'], {})
    pkg_rows.append([p['main_name'], p['sub_name'], '패키지', p['id'], dp.get('title') or p['title'],
                     p.get('period'), won(p.get('price')), won(p.get('list_price')),
                     pct(p.get('discount')), dp.get('track_count'),
                     BASE + 'package_detail.php?id=%s' % p['id']])
write_csv('itsdong_패키지목록.csv',
          ['대분류', '대표소분류', '구분', 'ID', '패키지명', '수강기간', '가격', '정가',
           '할인%', '포함트랙수', 'URL'], pkg_rows)

trk_rows = []
for p in pkgs_sorted:
    dp = pkgdet.get(p['id'], {})
    pname = dp.get('title') or p['title']
    for i, t in enumerate(dp.get('tracks', []), 1):
        trk_rows.append([p['id'], pname, i, t['id'], t['title'],
                         det(t['id']).get('lesson_count'),
                         BASE + 'lecture_detail.php?id=%s' % t['id']])
write_csv('itsdong_패키지구성.csv',
          ['패키지ID', '패키지명', '순번', '트랙ID', '트랙제목', '트랙차시수', '트랙URL'], trk_rows)

cur_rows = []
for l in sorted(lectures, key=lambda l: (MAIN_ORDER.index(l['main_name']) if l['main_name'] in MAIN_ORDER else 9,
                                         SUB_ORDER.get((l['main_name'], l['sub_name']), 999), l['title'])):
    d = det(l['id'])
    for les in d.get('lessons', []):
        cur_rows.append([l['main_name'], l['sub_name'], l['id'], l['title'],
                         les['no'], les['title'], les.get('time')])
write_csv('itsdong_차시목차.csv',
          ['대분류', '소분류', '강좌ID', '강좌제목', '차시', '차시제목', '시간'], cur_rows)

print('CSV 4종 저장 (강좌 %d, 패키지 %d, 트랙 %d, 차시 %d행)' % (
    len(lec_rows), len(pkg_rows), len(trk_rows), len(cur_rows)))
